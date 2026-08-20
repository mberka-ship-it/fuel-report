#!/usr/bin/env python3
"""Refresh official fuel datasets and inject them into the portable dashboard.

The dashboard is a self-contained Data Analytics artifact. Its interactive reader
loads a gzip/base64 JSON payload embedded in ``index.html``. This script updates
that payload in place so the public page changes when the workflow runs; updating
``data/snapshot.json`` alone would not change the visible dashboard.
"""

from __future__ import annotations

import base64
import calendar
import gzip
import html
import io
import json
import os
import re
import subprocess
import sys
from datetime import date, datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, unquote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INDEX_PATH = ROOT / "index.html"
SNAPSHOT_PATH = ROOT / "data" / "snapshot.json"

FUEL_PLAN_URL = "https://fuelplan.gov.au/fuel-statistics"
MSO_XLSX_URL = "https://www.dcceew.gov.au/sites/default/files/documents/mso-weekly-snapshot-timeseries.xlsx"
APS_PAGE_URL = "https://www.energy.gov.au/publications/australian-petroleum-statistics-2026"
USER_AGENT = "SEQ-fuel-security-monitor/2.0 (+https://github.com/mberka-ship-it/fuel-report)"

MONTHS = {name: index for index, name in enumerate(calendar.month_name) if name}
MONTHS_LOWER = {name.lower(): index for name, index in MONTHS.items()}


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        self.parts.append(data)


def fetch(url: str, *, attempts: int = 2) -> bytes:
    """Fetch a source with bounded retries and a descriptive failure.

    curl is preinstalled on GitHub's Ubuntu runners and handles these Australian
    Government endpoints more reliably than Python's urllib TLS stack.
    """
    command = [
        "curl", "--http1.1", "--fail", "--location", "--silent", "--show-error",
        "--retry", str(max(0, attempts - 1)), "--retry-all-errors",
        "--connect-timeout", "10", "--max-time", "45",
        "--user-agent", USER_AGENT, url,
    ]
    result = subprocess.run(command, check=False, capture_output=True)
    if result.returncode != 0:
        detail = result.stderr.decode("utf-8", errors="replace").strip()
        raise RuntimeError(f"Could not fetch {url}: {detail or f'curl exit {result.returncode}'}")
    return result.stdout


def html_text(content: bytes) -> str:
    parser = TextExtractor()
    parser.feed(content.decode("utf-8", errors="replace"))
    return " ".join(" ".join(parser.parts).replace("\xa0", " ").split())


def must_match(pattern: str, text: str, label: str, flags: int = 0) -> re.Match[str]:
    match = re.search(pattern, text, flags)
    if not match:
        raise ValueError(f"Could not parse {label}; source layout may have changed")
    return match


def number(value: str) -> float:
    return float(value.replace(",", ""))


def compact_number(value: float) -> int | float:
    rounded = round(value, 3)
    return int(rounded) if rounded.is_integer() else rounded


def pct(value: str) -> float:
    return number(value) / 100


def source_date(day: str, month: str, year: str) -> date:
    return date(int(year), MONTHS[month], int(day))


def date_label(value: date, *, year: bool = True) -> str:
    return f"{value.day} {calendar.month_name[value.month]}" + (f" {value.year}" if year else "")


def short_date_label(value: date) -> str:
    return f"{value.day} {calendar.month_abbr[value.month]}"


def month_label(value: date) -> str:
    return f"{calendar.month_name[value.month]} {value.year}"


def prior_month(value: date, months: int) -> date:
    serial = value.year * 12 + value.month - 1 - months
    return date(serial // 12, serial % 12 + 1, 1)


def movement_word(change: float, *, higher: str = "rose", lower: str = "fell") -> str:
    if change > 0:
        return higher
    if change < 0:
        return lower
    return "was unchanged"


def signed_int_from_parenthetical(value: str) -> int:
    value = value.strip().lower()
    if "no change" in value:
        return 0
    match = must_match(r"([+-]?\d+)", value, "stock-out weekly change")
    return int(match.group(1))


def parse_fuel_plan(content: bytes) -> dict[str, Any]:
    text = html_text(content)

    price_date_match = must_match(r"Retail price (\d{1,2}) ([A-Za-z]+) (\d{4})", text, "retail price date")
    stock_date_match = must_match(r"Days of fuel reserves held under MSO (\d{1,2}) ([A-Za-z]+) (\d{4})", text, "MSO date")
    flow_dates_match = must_match(
        r"Ships on water to Australia As at (\d{1,2}) ([A-Za-z]+) (\d{4}) As at (\d{1,2}) ([A-Za-z]+) (\d{4})",
        text,
        "shipment dates",
    )
    outage_date_match = must_match(r"Retail stock-outs[\s\S]*?Data as at (\d{1,2}) ([A-Za-z]+) (\d{4})", text, "retail availability date")

    price_date = source_date(*price_date_match.groups())
    stock_date = source_date(*stock_date_match.groups())
    flow_date = source_date(*flow_dates_match.groups()[:3])
    prior_flow_date = source_date(*flow_dates_match.groups()[3:])
    outage_date = source_date(*outage_date_match.groups())

    plan_level = int(must_match(r"Level (\d+) - National Fuel Security Plan", text, "plan level").group(1))
    prices = must_match(
        r"5 largest cities\* \$([\d.]+) \(([+-]?\d+)%\) \$([\d.]+) \(([+-]?\d+)%\).*?BRIS \$([\d.]+) \(([+-]?\d+)%\) \$([\d.]+) \(([+-]?\d+)%\)",
        text,
        "retail prices",
    ).groups()
    pre_conflict = must_match(
        r"pre-conflict levels.*?petrol prices.*?(\d+) cpl higher.*?diesel prices.*?(\d+) cpl higher",
        text,
        "pre-conflict retail change",
        re.IGNORECASE,
    ).groups()
    benchmarks = must_match(
        r"Brent Crude US\$([\d.]+) \(([+-]?\d+)%\) \+?([+-]?\d+)% Singapore Gasoil \(diesel\) US\$([\d.]+) \(([+-]?\d+)%\) \+?([+-]?\d+)%",
        text,
        "international benchmark prices",
    ).groups()
    cover = must_match(
        r"Petrol ([\d.]+) ([\d.]+) Diesel ([\d.]+) ([\d.]+) Jet fuel ([\d.]+) ([\d.]+)",
        text,
        "MSO cover days",
    ).groups()
    scheduled = number(must_match(r"At least ([\d.]+) billion litres", text, "forward orders").group(1))
    shipments = must_match(
        r"Crude oil (\d+) tankers equivalent to ([\d.]+) days (\d+) tankers equivalent to ([\d.]+) days Clean refined products (\d+) tankers equivalent to ([\d.]+) days (\d+) tankers equivalent to ([\d.]+) days",
        text,
        "ships on water",
    ).groups()
    qld = must_match(
        r"QLD \(([\d,]+) sites\) (\d+) \(([^)]*)\) (\d+) \(([^)]*)\) (\d+) \(([^)]*)\)",
        text,
        "Queensland stock-outs",
    ).groups()

    site_count = int(qld[0].replace(",", ""))
    petrol_only, diesel_only, both = int(qld[1]), int(qld[3]), int(qld[5])
    petrol_change = signed_int_from_parenthetical(qld[2]) + signed_int_from_parenthetical(qld[6])
    diesel_change = signed_int_from_parenthetical(qld[4]) + signed_int_from_parenthetical(qld[6])
    petrol_affected, diesel_affected = petrol_only + both, diesel_only + both
    petrol_prior, diesel_prior = petrol_affected - petrol_change, diesel_affected - diesel_change

    current_crude_tankers, current_crude_days, prior_crude_tankers, prior_crude_days = map(number, shipments[:4])
    current_clean_tankers, current_clean_days, prior_clean_tankers, prior_clean_days = map(number, shipments[4:])

    fuelplan_summary = [{
        "plan_level": plan_level,
        "brisbane_petrol": number(prices[4]),
        "brisbane_petrol_wow": pct(prices[5]),
        "brisbane_diesel": number(prices[6]),
        "brisbane_diesel_wow": pct(prices[7]),
        "qld_petrol_affected_share": petrol_affected / site_count,
        "qld_petrol_affected_prior_share": petrol_prior / site_count,
        "qld_diesel_affected_share": diesel_affected / site_count,
        "qld_diesel_affected_prior_share": diesel_prior / site_count,
        "clean_pipeline_days": compact_number(current_clean_days),
        "clean_pipeline_prior_days": compact_number(prior_clean_days),
        "forward_orders_bn_l": scheduled,
        "diesel_stock_days": compact_number(number(cover[2])),
        "diesel_stock_vs_march_days": compact_number(number(cover[2]) - number(cover[3])),
    }]
    qld_outages = [
        {
            "fuel": "Petrol",
            "unleaded_or_diesel_only": petrol_only,
            "both_out": both,
            "affected_sites": petrol_affected,
            "prior_affected_sites": petrol_prior,
            "weekly_change_sites": petrol_change,
            "affected_share": round(petrol_affected / site_count, 4),
        },
        {
            "fuel": "Diesel",
            "unleaded_or_diesel_only": diesel_only,
            "both_out": both,
            "affected_sites": diesel_affected,
            "prior_affected_sites": diesel_prior,
            "weekly_change_sites": diesel_change,
            "affected_share": round(diesel_affected / site_count, 4),
        },
    ]
    shipment_days = [
        {"flow": "Crude oil", "period": short_date_label(prior_flow_date), "days": compact_number(prior_crude_days), "tankers": compact_number(prior_crude_tankers)},
        {"flow": "Crude oil", "period": short_date_label(flow_date), "days": compact_number(current_crude_days), "tankers": compact_number(current_crude_tankers)},
        {"flow": "Clean products", "period": short_date_label(prior_flow_date), "days": compact_number(prior_clean_days), "tankers": compact_number(prior_clean_tankers)},
        {"flow": "Clean products", "period": short_date_label(flow_date), "days": compact_number(current_clean_days), "tankers": compact_number(current_clean_tankers)},
    ]
    price_movements = [
        {"market": "Five-city petrol", "current": number(prices[0]), "unit": "A$/L", "weekly_change": pct(prices[1]), "change_since_pre_conflict": number(pre_conflict[0]) / 100, "since_unit": "A$/L"},
        {"market": "Five-city diesel", "current": number(prices[2]), "unit": "A$/L", "weekly_change": pct(prices[3]), "change_since_pre_conflict": number(pre_conflict[1]) / 100, "since_unit": "A$/L"},
        {"market": "Brent crude", "current": number(benchmarks[0]), "unit": "US$/bbl", "weekly_change": pct(benchmarks[1]), "change_since_pre_conflict": number(benchmarks[2]), "since_unit": "%"},
        {"market": "Singapore gasoil", "current": number(benchmarks[3]), "unit": "US$/bbl", "weekly_change": pct(benchmarks[4]), "change_since_pre_conflict": number(benchmarks[5]), "since_unit": "%"},
    ]
    benchmark_change = [
        {"benchmark": "Brent crude", "change": pct(benchmarks[2]), "current_usd_bbl": number(benchmarks[0]), "weekly_change": pct(benchmarks[1])},
        {"benchmark": "Singapore gasoil", "change": pct(benchmarks[5]), "current_usd_bbl": number(benchmarks[3]), "weekly_change": pct(benchmarks[4])},
    ]

    return {
        "dates": {"prices": price_date, "stocks": stock_date, "flows": flow_date, "flows_prior": prior_flow_date, "outages": outage_date},
        "fuelplan_summary": fuelplan_summary,
        "qld_outages": qld_outages,
        "shipment_days": shipment_days,
        "price_movements": price_movements,
        "benchmark_change": benchmark_change,
        "site_count": site_count,
        "cover": {"Petrol": number(cover[0]), "Diesel": number(cover[2]), "Jet fuel": number(cover[4])},
        "cover_march": {"Petrol": number(cover[1]), "Diesel": number(cover[3]), "Jet fuel": number(cover[5])},
    }


def worksheet_rows(workbook: Any, sheet: str) -> list[dict[str, Any]]:
    ws = workbook[sheet]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def parse_mso(content: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]], date]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    config = {
        "Gasoline": ("Petrol", "Stocks required under MSO (ML)", "Stocks required under MSO after s16A Temporary Reductions applied [1]", "Stock held under MSO (ML)", "Stock held under MSO (Days equivalent) [2]"),
        "Diesel": ("Diesel", "Stocks required under MSO (ML)", "Stocks required under MSO after s16A Temporary Reductions applied [1]", "Stock held under MSO (ML)", "Stock held under MSO (Days equivalent) [2]"),
        "Kerosene": ("Jet fuel", "Stocks required under MSO (ML)", None, "Stock held under MSO (ML)", "Stock held under MSO (Days equivalent) [2]"),
    }
    trend: list[dict[str, Any]] = []
    latest: list[dict[str, Any]] = []
    latest_date: date | None = None
    for sheet, (fuel, baseline_col, effective_col, held_col, days_col) in config.items():
        parsed: list[dict[str, Any]] = []
        for row in worksheet_rows(workbook, sheet):
            obligation = row.get("Obligation Date")
            held = row.get(held_col)
            days = row.get(days_col)
            if not isinstance(obligation, datetime) or not isinstance(held, (int, float)) or not isinstance(days, (int, float)):
                continue
            baseline = row.get(baseline_col)
            effective = row.get(effective_col) if effective_col else baseline
            if not isinstance(baseline, (int, float)):
                raise ValueError(f"Missing baseline MSO requirement for {fuel}")
            if not isinstance(effective, (int, float)):
                effective = baseline
            parsed.append({
                "date": obligation.date(),
                "fuel": fuel,
                "held_ml": compact_number(float(held)),
                "days": compact_number(float(days)),
                "baseline_required_ml": compact_number(float(baseline)),
                "effective_required_ml": compact_number(float(effective)),
            })
        if not parsed:
            raise ValueError(f"No usable MSO observations in {sheet}")
        for row in parsed[-12:]:
            trend.append({"date": row["date"].isoformat(), "fuel": fuel, "days": row["days"], "held_ml": row["held_ml"]})
        row = parsed[-1]
        held_value = float(row["held_ml"])
        baseline_value = float(row["baseline_required_ml"])
        effective_value = float(row["effective_required_ml"])
        latest.append({
            "fuel": fuel,
            "as_at": row["date"].isoformat(),
            "held_ml": row["held_ml"],
            "days": row["days"],
            "baseline_required_ml": row["baseline_required_ml"],
            "effective_required_ml": row["effective_required_ml"],
            "surplus_vs_baseline": round(held_value / baseline_value - 1, 3),
            "surplus_vs_effective": round(held_value / effective_value - 1, 3),
        })
        latest_date = row["date"] if latest_date is None else min(latest_date, row["date"])
    trend.sort(key=lambda row: (row["date"], row["fuel"]))
    assert latest_date is not None
    return trend, latest, latest_date


def latest_aps_link(page: bytes) -> tuple[str, date]:
    document = page.decode("utf-8", errors="replace")
    candidates: list[tuple[date, str]] = []
    for href in re.findall(r'href=["\']([^"\']+\.xlsx(?:\?[^"\']*)?)["\']', document, re.IGNORECASE):
        decoded = unquote(href).lower().replace("-", "_")
        match = re.search(r"data[_ ]extract[_ ]([a-z]+)[_ ](20\d{2})", decoded)
        if not match or match.group(1) not in MONTHS_LOWER:
            continue
        month = MONTHS_LOWER[match.group(1)]
        candidates.append((date(int(match.group(2)), month, 1), urljoin(APS_PAGE_URL, html.unescape(href))))
    if not candidates:
        raise ValueError("Could not find an APS monthly data extract link")
    latest_date, link = max(candidates, key=lambda item: item[0])
    return link, latest_date


def parse_aps(content: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], date]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sales_rows = worksheet_rows(workbook, "Sales by state and territory")
    qld = [row for row in sales_rows if row.get("State") == "QLD" and isinstance(row.get("Month"), datetime)]
    if not qld:
        raise ValueError("No Queensland sales data found in APS workbook")
    qld.sort(key=lambda row: row["Month"])
    latest = qld[-1]["Month"].date().replace(day=1)
    baseline = prior_month(latest, 12)
    previous = prior_month(latest, 1)
    previous_year_end = baseline
    previous_year_start = prior_month(previous_year_end, 11)
    trailing_start = prior_month(latest, 11)
    product_cols = {
        "Petrol": "Automotive gasoline: total (ML)",
        "Diesel": "Diesel oil: total",
        "Jet fuel": "Aviation turbine fuel: total (ML)",
    }
    by_month = {row["Month"].date().replace(day=1): row for row in qld}
    required = [baseline, previous, latest]
    if any(month not in by_month for month in required):
        raise ValueError("APS workbook does not contain the months needed for comparisons")

    demand_index: list[dict[str, Any]] = []
    demand_summary: list[dict[str, Any]] = []
    window_months = sorted(month for month in by_month if baseline <= month <= latest)
    for product, column in product_cols.items():
        base_value = float(by_month[baseline][column])
        for month in window_months:
            value = float(by_month[month][column])
            demand_index.append({
                "month": month.isoformat(),
                "period": f"{calendar.month_abbr[month.month]} {month.year}",
                "product": product,
                "index": round(value / base_value * 100, 1),
                "volume_ml": round(value, 1),
            })
        current = float(by_month[latest][column])
        prior = float(by_month[previous][column])
        year_ago = float(by_month[baseline][column])
        trailing = sum(float(by_month[month][column]) for month in by_month if trailing_start <= month <= latest)
        prior_trailing = sum(float(by_month[month][column]) for month in by_month if previous_year_start <= month <= previous_year_end)
        demand_summary.append({
            "product": product,
            "latest_month_ml": round(current, 1),
            "mom_change": round(current / prior - 1, 3),
            "yoy_change": round(current / year_ago - 1, 3),
            "trailing_12m_ml": round(trailing, 1),
            "trailing_12m_change": round(trailing / prior_trailing - 1, 3),
        })

    import_rows = worksheet_rows(workbook, "Imports volume by country")
    import_months = [row["Month"].date().replace(day=1) for row in import_rows if isinstance(row.get("Month"), datetime)]
    import_latest = max(month for month in import_months if month <= latest)
    relevant = [row for row in import_rows if isinstance(row.get("Month"), datetime) and row["Month"].date().replace(day=1) == import_latest]
    columns = ["Automotive gasoline (ML)", "Diesel oil (ML)", "Aviation turbine fuel (ML)"]
    totals: list[tuple[str, float]] = []
    for row in relevant:
        volume = sum(float(row.get(column) or 0) for column in columns)
        if volume > 0:
            origin = str(row.get("Source country") or "Unknown")
            origin = origin.replace("Korea, Republic of (South)", "South Korea").replace("China (excludes SARs and Taiwan)", "China")
            totals.append((origin, volume))
    grand_total = sum(value for _, value in totals)
    if grand_total <= 0:
        raise ValueError("No major refined-product import volumes found in APS workbook")
    totals.sort(key=lambda item: item[1], reverse=True)
    top = totals[:7]
    other = grand_total - sum(value for _, value in top)
    import_origins = [{"origin": origin, "volume_ml": round(value, 1), "share": round(value / grand_total, 4)} for origin, value in top]
    if other > 0.05:
        import_origins.append({"origin": "Other", "volume_ml": round(other, 1), "share": round(other / grand_total, 4)})
    return demand_index, demand_summary, import_origins, latest


def decode_artifact(document: str) -> tuple[dict[str, Any], re.Match[str]]:
    pattern = re.compile(
        r'(<template\b[^>]*\bid=["\']data-analytics-portable-artifact-payload-source["\'][^>]*>)(.*?)(</template>)',
        re.IGNORECASE | re.DOTALL,
    )
    match = pattern.search(document)
    if not match:
        raise ValueError("Portable dashboard payload template was not found in index.html")
    encoded = re.sub(r"\s+", "", match.group(2))
    payload = gzip.decompress(base64.b64decode(encoded)).decode("utf-8")
    return json.loads(payload), match


def encode_artifact(artifact: dict[str, Any]) -> str:
    payload = json.dumps(artifact, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    encoded = base64.b64encode(gzip.compress(payload, compresslevel=9, mtime=0)).decode("ascii")
    return "\n".join(encoded[index:index + 100] for index in range(0, len(encoded), 100))


def snapshot_sql(datasets: dict[str, list[dict[str, Any]]]) -> str:
    selects: list[str] = []
    for dataset, rows in datasets.items():
        for row_number, row in enumerate(rows, start=1):
            for field, value in row.items():
                safe = lambda item: str(item).replace("'", "''")
                selects.append(f"SELECT '{safe(dataset)}' AS dataset, {row_number} AS row_number, '{safe(field)}' AS field, '{safe(value)}' AS value")
    return "\nUNION ALL\n".join(selects)


def find_by_id(items: list[dict[str, Any]], item_id: str) -> dict[str, Any]:
    for item in items:
        if item.get("id") == item_id:
            return item
    raise KeyError(f"Artifact item {item_id!r} not found")


def build_scenarios(mso_latest: list[dict[str, Any]]) -> list[dict[str, Any]]:
    starting = {row["fuel"]: float(row["days"]) for row in mso_latest}
    scenarios = [
        ("Base — 98% petrol / 96% diesel / 98% jet replenishment", {"Petrol": 0.98, "Diesel": 0.96, "Jet fuel": 0.98}),
        ("Tight — 90% / 85% / 88%", {"Petrol": 0.90, "Diesel": 0.85, "Jet fuel": 0.88}),
        ("Severe — 75% / 60% / 70%", {"Petrol": 0.75, "Diesel": 0.60, "Jet fuel": 0.70}),
    ]
    rows = []
    for label, replenishment in scenarios:
        week8 = {fuel: max(0, starting[fuel] + 8 * 7 * (ratio - 1)) for fuel, ratio in replenishment.items()}
        rows.append({"scenario": label, "petrol_days": round(week8["Petrol"], 1), "diesel_days": round(week8["Diesel"], 1), "jet_days": round(week8["Jet fuel"], 1)})
    return rows


def update_artifact(
    artifact: dict[str, Any],
    fuel: dict[str, Any],
    mso_trend: list[dict[str, Any]],
    mso_latest: list[dict[str, Any]],
    mso_date: date,
    demand_index: list[dict[str, Any]],
    demand_summary: list[dict[str, Any]],
    import_origins: list[dict[str, Any]],
    aps_date: date,
    aps_url: str,
) -> dict[str, Any]:
    generated = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    today = datetime.now(timezone.utc).date()
    datasets = artifact["snapshot"]["datasets"]
    datasets.update({
        "fuelplan_summary": fuel["fuelplan_summary"],
        "qld_outages": fuel["qld_outages"],
        "shipment_days": fuel["shipment_days"],
        "price_movements": fuel["price_movements"],
        "benchmark_change": fuel["benchmark_change"],
        "mso_trend": mso_trend,
        "mso_latest": mso_latest,
        "qld_demand_index": demand_index,
        "qld_demand_summary": demand_summary,
        "import_origins": import_origins,
        "scenario_week8": build_scenarios(mso_latest),
    })

    fuel_date = fuel["dates"]["flows"]
    price_date = fuel["dates"]["prices"]
    outage_date = fuel["dates"]["outages"]
    prior_flow_date = fuel["dates"]["flows_prior"]
    latest_summary = fuel["fuelplan_summary"][0]
    petrol_outage, diesel_outage = fuel["qld_outages"]
    current_crude = next(row for row in fuel["shipment_days"] if row["flow"] == "Crude oil" and row["period"] == short_date_label(fuel_date))
    prior_crude = next(row for row in fuel["shipment_days"] if row["flow"] == "Crude oil" and row["period"] == short_date_label(prior_flow_date))
    current_clean = next(row for row in fuel["shipment_days"] if row["flow"] == "Clean products" and row["period"] == short_date_label(fuel_date))
    prior_clean = next(row for row in fuel["shipment_days"] if row["flow"] == "Clean products" and row["period"] == short_date_label(prior_flow_date))
    latest_mso = {row["fuel"]: row for row in mso_latest}
    top_two_share = sum(row["share"] for row in import_origins[:2])

    # Replace automatically derived development rows, leaving reviewed geopolitical
    # and policy narrative intact.
    retained = [
        row for row in datasets.get("developments", [])
        if row.get("source") not in {"Fuel Plan"}
        and not (row.get("source") == "DCCEEW" and str(row.get("development", "")).startswith("MSO stocks were"))
    ]
    derived = [
        {
            "date": fuel_date.isoformat(),
            "development": f"Clean-product ships were {current_clean['tankers']} and {current_clean['days']} consumption-equivalent days; at least {latest_summary['forward_orders_bn_l']}bn L is scheduled over four weeks.",
            "implication": "Inbound cover remains material, but the weekly ship count is a rolling pipeline snapshot rather than guaranteed delivered stock.",
            "source": "Fuel Plan",
        },
        {
            "date": outage_date.isoformat(),
            "development": f"Queensland diesel-affected sites were {diesel_outage['affected_sites']} ({diesel_outage['weekly_change_sites']:+d} week on week); petrol-affected sites were {petrol_outage['affected_sites']} ({petrol_outage['weekly_change_sites']:+d}).",
            "implication": "This is a statewide availability proxy; it does not isolate Southeast Queensland or terminal inventories.",
            "source": "Fuel Plan",
        },
        {
            "date": mso_date.isoformat(),
            "development": f"MSO stocks were {latest_mso['Petrol']['days']} days petrol, {latest_mso['Diesel']['days']} days diesel and {latest_mso['Jet fuel']['days']} days jet fuel.",
            "implication": "These are government-mandated industry holdings, not a government-owned strategic reserve.",
            "source": "DCCEEW",
        },
    ]
    datasets["developments"] = sorted(derived + retained, key=lambda row: row.get("date", ""), reverse=True)

    manifest = artifact["manifest"]
    manifest["generatedAt"] = generated
    manifest["description"] = f"Latest Southeast Queensland signals in Australian and global context — checked {date_label(today)}. Each panel shows its source period."
    artifact["snapshot"]["generatedAt"] = generated

    find_by_id(manifest["cards"], "bris_petrol")["description"] = f"Brisbane average retail petrol price on {date_label(price_date)}."
    find_by_id(manifest["cards"], "bris_diesel")["description"] = f"Brisbane average retail diesel price on {date_label(price_date)}."

    mso_chart = find_by_id(manifest["charts"], "mso_days_trend")
    mso_chart["subtitle"] = f"Latest weekly MSO position through {date_label(mso_date)}; hover points for fuel-specific values."
    shipment_chart = find_by_id(manifest["charts"], "shipment_pipeline")
    shipment_chart["subtitle"] = f"Clean-product cover {movement_word(float(current_clean['days']) - float(prior_clean['days']))} from {prior_clean['days']} to {current_clean['days']} days; crude cover {movement_word(float(current_crude['days']) - float(prior_crude['days']))} from {prior_crude['days']} to {current_crude['days']} days."
    imports_chart = find_by_id(manifest["charts"], "import_origins")
    imports_chart["title"] = f"{month_label(aps_date)} major refined-product imports by origin"
    imports_chart["subtitle"] = f"{import_origins[0]['origin']} and {import_origins[1]['origin']} supplied {top_two_share:.1%} of petrol, diesel and jet imports in {calendar.month_name[aps_date.month]}."
    imports_chart["xAxisTitle"] = f"Share of {month_label(aps_date)} major refined-product imports"
    demand_chart = find_by_id(manifest["charts"], "qld_demand_index")
    demand_chart["subtitle"] = f"Monthly Queensland volumes through {month_label(aps_date)}; the date axis represents whole calendar months."
    demand_chart["headerMarkdown"] = f"Monthly product sales, indexed to {month_label(prior_month(aps_date, 12))} = 100; Queensland statewide, not SEQ-only."
    demand_chart["yAxisTitle"] = f"Index ({month_label(prior_month(aps_date, 12))} = 100)"
    demand_chart["encodings"]["x"].update({"field": "period", "type": "nominal", "label": "Reporting month"})

    qld_table = find_by_id(manifest["tables"], "qld_availability")
    qld_table["subtitle"] = f"Statewide reporting across {fuel['site_count']:,} sites as at {date_label(outage_date)}; changes are week on week."
    mso_table = find_by_id(manifest["tables"], "mso_stock_detail")
    mso_table["subtitle"] = f"Latest official weekly position as at {date_label(mso_date)}. Baseline is statutory; effective requirement includes any temporary reduction shown in the source workbook."
    demand_table = find_by_id(manifest["tables"], "qld_demand_summary")
    demand_table["subtitle"] = f"{month_label(aps_date)} and trailing-12-month movements from Australian Petroleum Statistics."
    demand_table["columns"][1]["field"] = "latest_month_ml"
    demand_table["columns"][1]["label"] = f"{calendar.month_abbr[aps_date.month]} {str(aps_date.year)[2:]}, ML"
    developments_table = find_by_id(manifest["tables"], "developments")
    developments_table["subtitle"] = f"Automatically refreshed quantitative releases plus reviewed policy and geopolitical context; checked {date_label(today)}."

    sources = manifest["sources"]
    aps_manifest_source = find_by_id(sources, "aps_june")
    aps_manifest_source.update({"label": f"Australian Petroleum Statistics — {month_label(aps_date)}", "path": aps_url, "href": aps_url})

    blocks = manifest["blocks"]
    find_by_id(blocks, "status_note")["body"] = (
        "## Current assessment\n\n"
        f"**Official data were checked on {date_label(today)}.** [Fuel Plan]({FUEL_PLAN_URL}) reports Plan level {latest_summary['plan_level']}, "
        f"Brisbane petrol at A${latest_summary['brisbane_petrol']:.2f}/L and diesel at A${latest_summary['brisbane_diesel']:.2f}/L. "
        f"The clean-product pipeline represents {current_clean['days']} consumption-equivalent days. "
        "These quantitative indicators refresh automatically; geopolitical interpretation remains a reviewed snapshot and is dated below."
    )
    find_by_id(blocks, "pipeline_heading")["body"] = (
        "## National buffer and shipment pipeline\n\n"
        f"The latest MSO release is dated **{date_label(mso_date)}** and the shipment snapshot **{date_label(fuel_date)}**. "
        "Stocks and cargoes are separate buffers: MSO data describe mandated industry holdings, while the tanker series describes a rolling inbound pipeline."
    )
    find_by_id(blocks, "market_heading")["body"] = (
        "## Market pressure and import exposure\n\n"
        f"Current weekly retail and benchmark prices come from [Fuel Plan]({FUEL_PLAN_URL}). "
        f"Country-of-origin and Queensland demand figures come from the latest published [Australian Petroleum Statistics]({aps_url}), covering **{month_label(aps_date)}**. "
        "Monthly dates are plotted at the first of the month for charting only; each point represents the whole month."
    )
    find_by_id(blocks, "limitations")["body"] = (
        "## Data limitations\n\n"
        f"- **Mixed reporting clocks:** prices run to {date_label(price_date)}, MSO stocks to {date_label(mso_date)}, ships to {date_label(fuel_date)}, retail availability to {date_label(outage_date)}, and petroleum trade/sales to {month_label(aps_date)}. A daily check cannot make weekly or monthly sources newer than their publishers.\n"
        "- **Monthly chart dates:** APS observations are stored at the first day of each month for plotting; for example, a 1 June point means the full June reporting month.\n"
        "- **No public SEQ stock ledger:** Queensland stock-outs and sales are statewide proxies; they cannot identify terminal inventories or suburb-level conditions.\n"
        "- **Pipeline aggregation:** public cargo data show tanker counts and consumption-equivalent days, not vessel names, ports, exact products, ETAs or cancellation risk.\n"
        "- **Stock ownership:** MSO volumes are held by regulated importers and refiners under government mandate. They are not the same as a government-owned strategic petroleum reserve.\n"
        "- **Narrative refresh:** quantitative official releases update automatically. Policy and geopolitical commentary is retained as a dated, reviewed snapshot rather than inferred unattended.\n"
        "- **Scenario model:** the eight-week sensitivity is mechanical and should not be treated as a probability-weighted forecast."
    )

    # Keep rich source provenance synchronized with the rows actually displayed.
    source_records = artifact.get("sources", [])
    source_datasets = {
        "fuelplan_latest": {key: datasets[key] for key in ["fuelplan_summary", "qld_outages", "shipment_days", "price_movements", "benchmark_change"]},
        "mso_weekly": {key: datasets[key] for key in ["mso_trend", "mso_latest"]},
        "aps_june": {key: datasets[key] for key in ["qld_demand_index", "qld_demand_summary", "import_origins"]},
        "scenario_method": {"scenario_week8": datasets["scenario_week8"]},
        "news_compilation": {"developments": datasets["developments"]},
    }
    for record in source_records:
        record_id = record.get("id")
        query = record.get("query")
        if record_id in source_datasets and isinstance(query, dict):
            query["executed_at"] = generated
            query["sql"] = snapshot_sql(source_datasets[record_id])
        if record_id == "aps_june":
            record.update({"label": f"Australian Petroleum Statistics — {month_label(aps_date)}", "path": aps_url, "href": aps_url})
            if isinstance(query, dict):
                query["url"] = aps_url
                query["filters"] = [f"Queensland sales through {month_label(aps_date)}", f"{month_label(aps_date)} source-country imports", "Major refined products = automotive gasoline + diesel + aviation turbine fuel"]
        elif record_id == "fuelplan_latest" and isinstance(query, dict):
            query["filters"] = [f"Prices to {date_label(price_date)}", f"Ships to {date_label(fuel_date)}", f"Queensland retail availability to {date_label(outage_date)}"]
        elif record_id == "mso_weekly" and isinstance(query, dict):
            query["filters"] = [f"Latest twelve weekly observations through {date_label(mso_date)}"]

    return {
        "generatedAt": generated,
        "sourcePeriods": {
            "fuelPlanPrices": price_date.isoformat(),
            "fuelPlanStocks": fuel["dates"]["stocks"].isoformat(),
            "fuelPlanFlows": fuel_date.isoformat(),
            "fuelPlanRetailAvailability": outage_date.isoformat(),
            "msoWeekly": mso_date.isoformat(),
            "apsMonthly": aps_date.isoformat(),
        },
        "sourceUrls": {"fuelPlan": FUEL_PLAN_URL, "mso": MSO_XLSX_URL, "aps": aps_url},
        "datasets": datasets,
    }


def main() -> int:
    fixture_dir_value = os.environ.get("FUEL_REFRESH_FIXTURE_DIR")
    if fixture_dir_value:
        fixture_dir = Path(fixture_dir_value)
        fuel_content = (fixture_dir / "fuelplan.html").read_bytes()
        aps_page = (fixture_dir / "aps.html").read_bytes()
        aps_url, advertised_aps_date = latest_aps_link(aps_page)
        mso_content = (fixture_dir / "mso.xlsx").read_bytes()
        aps_content = (fixture_dir / "aps.xlsx").read_bytes()
    else:
        fuel_content = fetch(FUEL_PLAN_URL)
        aps_page = fetch(APS_PAGE_URL)
        aps_url, advertised_aps_date = latest_aps_link(aps_page)
        mso_content = fetch(MSO_XLSX_URL)
        aps_content = fetch(aps_url)

    fuel = parse_fuel_plan(fuel_content)
    mso_trend, mso_latest, mso_date = parse_mso(mso_content)
    demand_index, demand_summary, import_origins, aps_date = parse_aps(aps_content)
    if aps_date != advertised_aps_date:
        raise ValueError(f"APS page advertises {advertised_aps_date}, but workbook ends at {aps_date}")

    document = INDEX_PATH.read_text(encoding="utf-8")
    artifact, payload_match = decode_artifact(document)
    public_snapshot = update_artifact(artifact, fuel, mso_trend, mso_latest, mso_date, demand_index, demand_summary, import_origins, aps_date, aps_url)

    encoded = encode_artifact(artifact)
    document = document[:payload_match.start(2)] + "\n" + encoded + "\n" + document[payload_match.end(2):]
    INDEX_PATH.write_text(document, encoding="utf-8")
    SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
    SNAPSHOT_PATH.write_text(json.dumps(public_snapshot, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Dashboard checked at {public_snapshot['generatedAt']}")
    for source, period in public_snapshot["sourcePeriods"].items():
        print(f"  {source}: {period}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"refresh failed: {error}", file=sys.stderr)
        raise
